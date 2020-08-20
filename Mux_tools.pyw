# -*- coding: utf-8 -*-
"""
开发者：周梦雄
最后更新日期：2020/8/20
"""
import sys
import os
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QMessageBox,
    QTableWidgetItem,
    QAbstractItemView,
)
from Ui_mux_tools import *
from PyQt5.QtCore import QDateTime
import sqlite3
from openpyxl import Workbook
import configparser
from configuration_databases import *
from datetime import datetime
from PyQt5.QtCore import QRegExp, QThread, pyqtSignal
from PyQt5.QtGui import QRegExpValidator
from urllib.request import urlopen
from urllib.error import URLError


# 重写三相线程类


class ThreadThree(QThread):
    signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def run(self):
        ex.session_add_iii()
        session_chip_id.commit()
        self.signal.emit("恭喜，上传成功！")


# 重写II采线程类


class ThreadTwo(QThread):
    signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()

    def run(self):
        ex.session_add_ii()
        session_chip_id.commit()
        self.signal.emit("恭喜，上传成功！")


# noinspection SqlDialectInspection,DuplicatedCode,SqlNoDataSourceInspection


class MyMainWindow(QMainWindow, Ui_STA_database_query):
    # noinspection SqlNoDataSourceInspection
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 审批单号、派工单号加入长度限制
        reg = QRegExp("[0-9]+$")
        approvalValidator = QRegExpValidator(reg)
        self.le_approval_no_iii.setValidator(approvalValidator)
        self.le_approval_no_ii.setValidator(approvalValidator)
        # 用来保存ID对应表
        self.result_unique_iii = []
        self.result_unique_ii = []
        # 设置默认产品型态
        self.cb_prod_type_III.setCurrentIndex(1)
        self.cb_prod_type_II.setCurrentIndex(4)
        # 数据库查询combobox并加载到应用中
        software_version_all = session.query(
            SoftwareVersion.software_version).all()
        customer_version_all = session.query(
            CustomerVersion.customer_version).all()
        vendor_code_all = session.query(VendorCode.vendor_code).all()
        software_date_all = session.query(SoftwareDate.software_date).all()
        software_version_all_list = [e[0] for e in software_version_all]
        customer_version_all_list = [e[0] for e in customer_version_all]
        vendor_code_all_list = [e[0] for e in vendor_code_all]
        software_date_all_list = [e[0] for e in software_date_all]
        self.cb_version_sw_II.addItems(software_version_all_list)
        self.cb_vendor_code_II.addItems(vendor_code_all_list)
        self.cb_date_sw_II.addItems(software_date_all_list)
        self.cb_ext_version_II.addItems(customer_version_all_list)
        self.cb_version_sw_III.addItems(software_version_all_list)
        self.cb_vendor_code_III.addItems(vendor_code_all_list)
        self.cb_date_sw_III.addItems(software_date_all_list)
        self.cb_ext_version_III.addItems(customer_version_all_list)

        # 实例化三相线程
        self.thread_iii = ThreadThree()
        self.thread_iii.signal.connect(
            self.finish_iii)  # 连接线程类中自定义信号槽到本类的自定义槽函数

        # 实例化II采线程
        self.thread_ii = ThreadTwo()
        self.thread_ii.signal.connect(self.finish_ii)  # 连接线程类中自定义信号槽到本类的自定义槽函数

        # 设置派工单默认前缀
        datetime_now_str = str(datetime.now())
        self.value_order_II.setText(
            'X' + datetime_now_str[:4] + datetime_now_str[5:7])
        self.value_order_III.setText(
            'X' + datetime_now_str[:4] + datetime_now_str[5:7])
        # 三相初始化
        self.tableWidget_III.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget_III.setColumnCount(3)
        self.tableWidget_III.setRowCount(99)
        self.tableWidget_III.resizeColumnsToContents()
        self.tableWidget_III.resizeRowsToContents()
        self.value_start_datetime_III.setDateTime(QDateTime.currentDateTime())
        self.textBrowser_III.append(
            "注意：3105芯片代码03；3911集中器芯片代码00，STA 01；北京、浙江集中器白名单关闭；"
        )
        self.textBrowser_III.setStyleSheet("* { color: #0000FF;}")
        self.statusbar.setStyleSheet(
            "* { color: #FF6666;font-size:30px;font-weight:bold;}"
        )
        self.listWidget.setSpacing(7)
        # 数据库路径
        # 生产路径
        # db_file_III = r"C:\Users\Dream\Desktop\更新带屏读ID软件\最新读ID(带屏)Debug  20190621\Debug带瓶的\MyProtocol.db"
        # 测试路径
        db_file_III = r"MyProtocol_iii.db"
        # 创建数据库连接对象
        self.conn_III = sqlite3.connect(db_file_III)
        # 创建游标对象
        self.cur_III = self.conn_III.cursor()
        self.start_date_III = self.value_start_datetime_III.dateTime().toString("yyyy-MM-dd HH:mm")
        self.sqlstring_III = r"SELECT ChipID,ModID,TTime FROM DataBackUp where ChipID<>'' and TTime>=? order by TTime asc;"
        self.btn_id_query_III.clicked.connect(self.click_query_iii)
        self.value_start_datetime_III.dateTimeChanged.connect(
            self.on_datetime_changed_iii)
        self.btn_export_id_III.clicked.connect(self.export_id_to_excel_iii)
        self.btn_cfg_query_III.clicked.connect(self.nv_query_iii)
        self.btn_save_III.clicked.connect(self.write_ini_iii)
        self.btn_upload_iii.clicked.connect(self.upload_iii)
        self.btn_upload_ii.clicked.connect(self.upload_ii)

        # II采初始化
        self.tableWidget_II.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableWidget_II.setColumnCount(3)
        self.tableWidget_II.setRowCount(99)
        self.tableWidget_II.resizeColumnsToContents()
        self.tableWidget_II.resizeRowsToContents()
        self.value_start_datetime_II.setDateTime(QDateTime.currentDateTime())
        self.textBrowser_II.append(
            "注意：3105芯片代码03；3911集中器芯片代码00，STA 01；北京、浙江集中器白名单关闭；"
        )
        self.textBrowser_II.setStyleSheet("* { color: #0000FF;}")
        self.statusbar.setStyleSheet(
            "* { color: #FF6666;font-size:30px;font-weight:bold;}"
        )
        # 数据库路径
        # 生产路径
        # db_file_II = r"C:\Users\Dream\Desktop\II采同时写模块ID和逻辑地址\Debug\MyProtocol.db"
        # 测试路径
        db_file_II = r"MyProtocol_ii.db"
        # 创建数据库连接对象
        self.conn_II = sqlite3.connect(db_file_II)
        # 创建游标对象
        self.cur_II = self.conn_II.cursor()
        self.start_date_II = self.value_start_datetime_II.dateTime().toString("yyyy-MM-dd HH:mm")
        self.sqlstring_II = r"""SELECT ChipIDRead, AssetIDWrite, datetime(datetime) AS test_datetime FROM (
           SELECT ChipIDRead,AssetIDWrite,substr(sTime, 1, 4) || '-' || substr(sTime, 5, 2) || '-' || substr(sTime, 7, 2) || 
           ' ' || substr(sTime, 10, 2) || ':' || substr(sTime, 12, 2) || ':' || substr(sTime, 14, 2) AS datetime
             FROM DataBackUp)  where ChipIDRead<>'' and test_datetime >= ? order by test_datetime asc;"""
        self.btn_id_query_II.clicked.connect(self.click_query_ii)
        self.value_start_datetime_II.dateTimeChanged.connect(
            self.on_datetime_changed_ii)
        self.btn_export_id_II.clicked.connect(self.export_id_to_excel_ii)
        self.btn_cfg_query_II.clicked.connect(self.nv_query_ii)
        self.btn_save_II.clicked.connect(self.write_ini_ii)

        # 添加参数到数据库信号与槽连接
        self.btn_software_version.clicked.connect(self.add_software_version)
        self.btn_customer_version.clicked.connect(self.add_customer_version)
        self.btn_vendor_code.clicked.connect(self.add_vendor_code)
        self.btn_software_date.clicked.connect(self.add_date)
        self.le_approval_no_iii.editingFinished.connect(
            self.approval_check_iii)
        self.le_approval_no_ii.editingFinished.connect(self.approval_check_ii)

        self.show()

    # 检查II采是否输入6位审批单号
    def approval_check_ii(self):
        if len(self.le_approval_no_ii.text()) != 6:
            QMessageBox.warning(
                self, '错误：', '您输入的审批单号不是6位，请重新输入！', QMessageBox.Ok)
            self.le_approval_no_ii.setFocus()

    # 检查三相是否输入6位审批单号
    def approval_check_iii(self):
        if len(self.le_approval_no_iii.text()) != 6:
            QMessageBox.warning(
                self, '错误：', '您输入的审批单号不是6位，请重新输入！', QMessageBox.Ok)
            self.le_approval_no_iii.setFocus()

    def on_datetime_changed_iii(self):
        self.start_date_III = self.value_start_datetime_III.dateTime().toString("yyyy-MM-dd HH:mm")

    def click_query_iii(self):
        self.tableWidget_III.clearContents()  # 每一次查询时清除表格中信息
        # 执行查询（传递开始测试日期时间参数）
        self.cur_III.execute(self.sqlstring_III, (self.start_date_III,))
        result_temp = self.cur_III.fetchall()
        result_id = [(r[0], r[1]) for r in result_temp]
        for i in result_id:
            if i not in self.result_unique_iii:
                self.result_unique_iii.append(i)
        try:
            # 自动设置ID倒数5个字符
            self.value_id_III.setText(result_temp[0][0][-5:])
            for k, i in enumerate(result_temp):
                print("----------", i)
                for w, j in enumerate(i):
                    if type(j) != str:
                        newItem = QTableWidgetItem(str(j))
                    else:
                        newItem = QTableWidgetItem(j)
                    # 根据循环标签一次对table中的格子进行设置
                    self.tableWidget_III.setItem(k, w, newItem)
            self.tableWidget_III.resizeColumnsToContents()
            self.tableWidget_III.resizeRowsToContents()
            self.textBrowser_III.setText("")
            self.textBrowser_III.append(
                "SELECT ChipID,ModID,TTime FROM DataBackUp where ChipID<>'' and TTime>=%r order by TTime asc;"
                % self.start_date_III
            )
            print("query button pressed")
        except IndexError:  # 查询记录无时，引发IndexError('list index out of range')异常。
            self.result_unique_iii = []
            QMessageBox.warning(self, '提示：', '查询结果为空！', QMessageBox.Ok)

    def export_id_to_excel_iii(self):
        if not self.result_unique_iii:
            QMessageBox.warning(
                self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
            return
        if len(self.value_order_III.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_III.setFocus()
            return
        else:
            wo = self.value_order_III.text().strip().upper()
            if self.cb_prod_type_III.currentText() == "通信单元（集中器I型/HPLC）":
                wo1 = wo + "-" + \
                      self.cb_prod_type_III.currentText().split(
                          '/')[0][-5:] + ".xlsx"
            elif self.cb_prod_type_III.currentText() == "II型采集器（HPLC）":
                wo1 = wo + "-" + \
                      self.cb_prod_type_III.currentText().split('（')[0] + ".xlsx"
            else:
                wo1 = wo + "-" + \
                      self.cb_prod_type_III.currentText().split(
                          '/')[0][-2:] + ".xlsx"
            # # 生产环境工作簿保存路径
            # path_name = os.path.join(
            #     r"C:\Users\Dream\Desktop\ID清单，请手下留情，勿删！！！", wo1)
            # 测试环境工作部保存路径
            path_name = os.path.join(os.getcwd(), wo1)
            # 新建工作簿
            wb = Workbook(path_name)
            ws = wb.create_sheet(wo, 0)
            ws.append(["芯片ID", "模块ID"])
            # # 查询结果
            # self.cur_III.execute(self.sqlstring_III, (self.start_date_III,))
            # result = self.cur_III.fetchall()
            for row in self.result_unique_iii:
                ws.append(list(row))
            self.statusbar.showMessage(
                "本批测试 %s 个模块，请注意检查是否有漏测！" % len(self.result_unique_iii))
            if self.result_unique_iii[0][0][-5:] != self.value_id_III.text().upper():
                self.statusbar.clearMessage()
                QMessageBox.warning(
                    self, "警告！", "你的首个ID不正确，请排查原因！", QMessageBox.Ok)
            else:
                wb.save(path_name)
                wb.close()
                QMessageBox.information(
                    self, "好消息！", "ID对应表已成功导出到excel表格！请核对左下角状态栏信息！", QMessageBox.Ok
                )

    def nv_query_iii(self):
        # 初始化
        conf_iii = configparser.ConfigParser()
        # 配置文件的绝对路径
        # 生产路径
        # conf_path_iii = r"C:\Users\Dream\Desktop\更新带屏读ID软件\最新读ID(带屏)Debug  20190621\Debug带瓶的\IniFile\FiterParam.ini"
        # 测试路径
        conf_path_iii = r"FiterParam.ini"
        # 读取配置文件
        conf_iii.read(conf_path_iii)
        # 返回section中option的值
        nv_configure_iii = "软件版本：%s 芯片代码：%s 版本日期：%s 外部版本：%s 厂商代码：%s" % (
            conf_iii.get("ErJiBiDui", "Value1"), conf_iii.get(
                "ErJiBiDui", "Value2"),
            conf_iii.get("ErJiBiDui", "Value3"), conf_iii.get(
                "ErJiBiDui", "Value4"),
            conf_iii.get("ErJiBiDui", "Value5"))
        self.textBrowser_III.setText("")
        self.textBrowser_III.append(nv_configure_iii)

    def write_ini_iii(self):
        config = configparser.ConfigParser()
        # 生产路径
        # path_name = r"C:\Users\Dream\Desktop\更新带屏读ID软件\最新读ID(带屏)Debug  20190621\Debug带瓶的\IniFile\FiterParam.ini"
        # 测试路径
        path_name = r"FiterParam.ini"
        config.read(path_name)  # 读文件
        section = r"ErJiBiDui"
        # 新增/修改配置文件的键值
        if len(self.cb_version_sw_III.currentText()) == 14:
            config.set(section, 'Value1',
                       self.cb_version_sw_III.currentText()[0:11] + '200')
        else:
            config.set(section, 'Value1', self.cb_version_sw_III.currentText())
        config.set(section, 'Value2', self.cb_chipcode_III.currentText())
        config.set(section, 'Value3', self.cb_date_sw_III.currentText())
        config.set(section, 'Value4', (self.cb_ext_version_III.currentText()[
                                       2:] + self.cb_ext_version_III.currentText()[:2]))
        config.set(section, 'Value5', self.cb_vendor_code_III.currentText())
        with open(path_name, 'w') as configfile:
            config.write(configfile)
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage("配置文件修改成功！", 3000)

    def upload_iii(self):
        if len(self.le_approval_no_iii.text()) == 0:
            QMessageBox.warning(
                self, '错误：', '您未输入审批单号，请输入！', QMessageBox.Ok)
            self.le_approval_no_iii.setFocus()
            return
        if len(self.value_order_III.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_III.setFocus()
            return
        if not self.result_unique_iii:
            QMessageBox.warning(
                self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
        else:
            try:
                # 判定服务器连接状态
                urlopen("https://www.baidu.com")
            except URLError:
                QMessageBox.warning(
                    self, '警告：', '无法连接到数据库，请检查网络连接状态！', QMessageBox.Ok)
            else:
                self.work_iii()

    def on_datetime_changed_ii(self):
        self.start_date_II = self.value_start_datetime_II.dateTime().toString("yyyy-MM-dd HH:mm")

    def click_query_ii(self):
        self.tableWidget_II.clearContents()  # 每一次查询时清除表格中信息
        # 查询结果
        self.cur_II.execute(self.sqlstring_II, (self.start_date_II,))
        result_temp = self.cur_II.fetchall()
        result_id = [(r[0], r[1]) for r in result_temp]
        for i in result_id:
            if i not in self.result_unique_ii:
                self.result_unique_ii.append(i)
        try:
            # 自动设置ID倒数5个字符
            self.value_id_II.setText(result_temp[0][0][-5:])
            for k, i in enumerate(result_temp):
                print("----------", i)
                for w, j in enumerate(i):
                    if type(j) != str:
                        newItem = QTableWidgetItem(str(j))
                    else:
                        newItem = QTableWidgetItem(j)
                    # 根据循环标签一次对table中的格子进行设置
                    self.tableWidget_II.setItem(k, w, newItem)
            self.tableWidget_II.resizeColumnsToContents()
            self.tableWidget_II.resizeRowsToContents()
            self.textBrowser_II.setText("")
            self.textBrowser_II.append(
                "SELECT ChipIDRead,AssetIDWrite,test_datetime FROM DataBackUp where ChipIDRead<>'' and ...;")
            print("find button pressed")
        except IndexError:
            self.result_unique_ii = []
            QMessageBox.warning(self, '提示：', '查询结果为空！', QMessageBox.Ok)

    def export_id_to_excel_ii(self):
        if not self.result_unique_ii:
            QMessageBox.warning(
                self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
            return
        if len(self.value_order_II.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_II.setFocus()
            return
        else:
            wo = self.value_order_II.text().strip().upper()
            if self.cb_prod_type_II.currentText() == "通信单元（集中器I型/HPLC）":
                wo1 = wo + "-" + \
                      self.cb_prod_type_II.currentText().split(
                          '/')[0][-5:] + ".xlsx"
            elif self.cb_prod_type_II.currentText() == "II型采集器（HPLC）":
                wo1 = wo + "-" + \
                      self.cb_prod_type_II.currentText().split('（')[0] + ".xlsx"
            else:
                wo1 = wo + "-" + \
                      self.cb_prod_type_II.currentText().split(
                          '/')[0][-2:] + ".xlsx"
            # 生产环境工作簿保存路径
            # path_name = os.path.join(
            #     r"C:\Users\Dream\Desktop\ID清单，请手下留情，勿删！！！", wo1)
            # 测试环境工作部保存路径
            path_name = os.path.join(os.getcwd(), wo1)
            # 新建工作簿
            wb = Workbook(path_name)
            ws = wb.create_sheet(wo, 0)
            ws.append(["芯片ID", "模块ID"])
            for row in self.result_unique_ii:
                ws.append(list(row))
            self.statusbar.showMessage(
                "本批测试 %s 个模块，请注意检查是否有漏测！" % len(self.result_unique_ii))
            if self.result_unique_ii[0][0][-5:] != self.value_id_II.text().upper():
                self.statusbar.clearMessage()
                QMessageBox.warning(
                    self, "警告！", "你的首个ID不正确，请排查原因！", QMessageBox.Ok)
            else:
                wb.save(path_name)
                QMessageBox.information(
                    self, "好消息！", "ID对应表已成功导出到excel表格！请核对左下角状态栏信息！", QMessageBox.Ok
                )

    def nv_query_ii(self):
        # 初始化
        conf_II = configparser.ConfigParser()
        # 配置文件的绝对路径
        # 生产路径
        # conf_path_II = r"C:\Users\Dream\Desktop\II采同时写模块ID和逻辑地址\Debug\IniFile\FiterParam.ini"
        # 测试路径
        conf_path_II = r"FiterParam.ini"
        # 读取配置文件
        conf_II.read(conf_path_II)
        # 返回section中option的值
        NV_configure_II = "软件版本：%s 芯片代码：%s 版本日期：%s 外部版本：%s 厂商代码：%s" % (
            conf_II.get("ErJiBiDui", "Value1"), conf_II.get(
                "ErJiBiDui", "Value2"),
            conf_II.get("ErJiBiDui", "Value3"), conf_II.get(
                "ErJiBiDui", "Value4"),
            conf_II.get("ErJiBiDui", "Value5"))
        self.textBrowser_II.setText("")
        self.textBrowser_II.append(NV_configure_II)

    def write_ini_ii(self):
        config = configparser.ConfigParser()
        # 生产路径
        # path_name = r"C:\Users\Dream\Desktop\II采同时写模块ID和逻辑地址\Debug\IniFile\FiterParam.ini"
        # 测试路径
        path_name = r"FiterParam.ini"
        config.read(path_name)  # 读文件
        section = r"ErJiBiDui"
        # 新增/修改配置文件的键值
        if len(self.cb_version_sw_II.currentText()) == 14:
            config.set(section, 'Value1',
                       self.cb_version_sw_II.currentText()[0:11] + '400')
        else:
            config.set(section, 'Value1', self.cb_version_sw_II.currentText())
        config.set(section, 'Value2', self.cb_chipcode_II.currentText())
        config.set(section, 'Value3', self.cb_date_sw_II.currentText())
        config.set(section, 'Value4', (self.cb_ext_version_II.currentText()[
                                       2:] + self.cb_ext_version_II.currentText()[:2]))
        config.set(section, 'Value5', self.cb_vendor_code_II.currentText())
        with open(path_name, 'w') as configfile:
            config.write(configfile)
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage("配置文件修改成功！", 3000)

    def upload_ii(self):
        if len(self.le_approval_no_ii.text()) == 0:
            QMessageBox.warning(
                self, '错误：', '您未输入审批单号，请输入！', QMessageBox.Ok)
            self.le_approval_no_ii.setFocus()
            return
        if len(self.value_order_II.text().strip()) <= 10:
            QMessageBox.warning(
                self, '错误：', '派工单号输入有误，请重新输入！', QMessageBox.Ok)
            self.value_order_II.setFocus()
            return
        if not self.result_unique_ii:
            QMessageBox.warning(
                self, '提示：', '请先点击查询按钮确认ID对应表有无异常！', QMessageBox.Ok)
        else:
            try:
                # 判定服务器连接状态
                urlopen("https://www.baidu.com")
            except URLError:
                QMessageBox.warning(
                    self, '警告：', '无法连接到数据库，请检查网络连接状态！', QMessageBox.Ok)
            else:
                self.work_ii()

    def closeEvent(self, event):  # 函数名固定不可变
        reply = QtWidgets.QMessageBox.question(
            self, u'警告', u'确认退出?', QtWidgets.QMessageBox.Yes, QtWidgets.QMessageBox.No)
        # QtWidgets.QMessageBox.question(self,u'弹窗名',u'弹窗内容',选项1,选项2)
        if reply == QtWidgets.QMessageBox.Yes:
            session.close()
            session_chip_id.close()
            self.conn_III.commit()
            self.cur_III.close()
            self.conn_III.close()
            self.conn_II.commit()
            self.cur_II.close()
            self.conn_II.close()
            event.accept()  # 关闭窗口
        else:
            event.ignore()  # 忽视点击X事件

    def add_software_version(self):
        if self.le_software_version.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "软件版本不能为空！", QMessageBox.Ok)
        elif session.query(SoftwareVersion.software_version).filter_by(
                software_version=self.le_software_version.text().strip().upper()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_version_sw_II.addItem(
                self.le_software_version.text().strip().upper())
            self.cb_version_sw_III.addItem(
                self.le_software_version.text().strip().upper())
            new_software_version = SoftwareVersion(
                software_version=self.le_software_version.text().strip().upper())
            session.add(new_software_version)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("软件版本添加成功！", 3000)

    def add_customer_version(self):
        if self.le_customer_version.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "外部版本不能为空！", QMessageBox.Ok)
        elif session.query(CustomerVersion.customer_version).filter_by(
                customer_version=self.le_customer_version.text().strip()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_ext_version_II.addItem(
                self.le_customer_version.text().strip())
            self.cb_ext_version_III.addItem(
                self.le_customer_version.text().strip())
            new_customer_version = CustomerVersion(
                customer_version=self.le_customer_version.text().strip())
            session.add(new_customer_version)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("外部版本添加成功！", 3000)

    def add_vendor_code(self):
        if self.le_vendor_code.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "厂商代码不能为空！", QMessageBox.Ok)
        elif session.query(VendorCode.vendor_code).filter_by(
                vendor_code=self.le_vendor_code.text().strip().upper()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_vendor_code_II.addItem(
                self.le_vendor_code.text().strip().upper())
            self.cb_vendor_code_III.addItem(
                self.le_vendor_code.text().strip().upper())
            new_vendor_code = VendorCode(
                vendor_code=self.le_vendor_code.text().strip().upper())
            session.add(new_vendor_code)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("厂商代码添加成功！", 3000)

    def add_date(self):
        if self.le_date.text() == '':
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "版本日期不能为空！", QMessageBox.Ok)
        elif session.query(SoftwareDate.software_date).filter_by(software_date=self.le_date.text().strip()).first():
            self.statusbar.clearMessage()
            QMessageBox.warning(
                self, "警告！", "数据库中已存在该数据，请勿重复添加！", QMessageBox.Ok)
        else:
            self.cb_date_sw_II.addItem(self.le_date.text().strip())
            self.cb_date_sw_III.addItem(self.le_date.text().strip())
            new_software_date = SoftwareDate(
                software_date=self.le_date.text().strip())
            session.add(new_software_date)
            session.commit()
            self.statusbar.setStyleSheet(
                "* { color: #00CD00;font-size:30px;font-weight:bold;}")
            self.statusbar.showMessage("版本日期添加成功！", 3000)

    # 定义三相线程启动函数
    def work_iii(self):
        self.thread_iii.start()
        self.btn_upload_iii.setDisabled(True)

    # 定义II采线程启动函数
    def work_ii(self):
        self.thread_ii.start()
        self.btn_upload_ii.setDisabled(True)

    # 定义三相线程完成后槽函数
    def finish_iii(self, info):
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage(info, 3000)
        self.btn_upload_iii.setEnabled(True)

    # 定义II采线程完成后槽函数
    def finish_ii(self, info):
        self.statusbar.setStyleSheet(
            "* { color: #00CD00;font-size:30px;font-weight:bold;}")
        self.statusbar.showMessage(info, 3000)
        self.btn_upload_ii.setEnabled(True)

    # 三相chip_id表数据添加到会话
    def session_add_iii(self):
        for tup in self.result_unique_iii:
            if tup[1] and not session_chip_id.query(ChipId).filter_by(
                    chip_id=tup[0]).first() and not session_chip_id.query(ChipId).filter_by(asset_no=tup[1]).first():
                new_work_order_no = WorkOrderNo(
                    work_order_no=self.value_order_III.text().strip().upper())
                new_approval_no = ApprovalNo(
                    approval_no=self.le_approval_no_iii.text())
                new_product_category = ProductCategory(
                    product_category=self.cb_prod_type_III.currentText())
                if not session_chip_id.query(WorkOrderNo).filter_by(
                        work_order_no=self.value_order_III.text().strip().upper()).first():
                    session_chip_id.add(new_work_order_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ApprovalNo).filter_by(
                        approval_no=self.le_approval_no_iii.text()).first():
                    session_chip_id.add(new_approval_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ProductCategory).filter_by(
                        product_category=self.cb_prod_type_III.currentText()).first():
                    session_chip_id.add(new_product_category)
                    session_chip_id.commit()
                row = ChipId(chip_id=tup[0], asset_no=tup[1],
                             work_order_no_id=session_chip_id.query(WorkOrderNo).filter_by(
                                 work_order_no=self.value_order_III.text().strip().upper()).first().id,
                             approval_no_id=session_chip_id.query(ApprovalNo).filter_by(
                                 approval_no=self.le_approval_no_iii.text()).first().id,
                             product_category_id=session_chip_id.query(ProductCategory).filter_by(
                                 product_category=self.cb_prod_type_III.currentText()).first().id)
                session_chip_id.add(row)

    # II采chip_id表数据添加到会话
    def session_add_ii(self):
        for tup in self.result_unique_ii:
            if tup[1] and not session_chip_id.query(ChipId).filter_by(
                    chip_id=tup[0]).first() and not session_chip_id.query(ChipId).filter_by(asset_no=tup[1]).first():
                new_work_order_no = WorkOrderNo(
                    work_order_no=self.value_order_II.text().strip().upper())
                new_approval_no = ApprovalNo(
                    approval_no=self.le_approval_no_ii.text())
                new_product_category = ProductCategory(
                    product_category=self.cb_prod_type_II.currentText())
                if not session_chip_id.query(WorkOrderNo).filter_by(
                        work_order_no=self.value_order_II.text().strip().upper()).first():
                    session_chip_id.add(new_work_order_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ApprovalNo).filter_by(
                        approval_no=self.le_approval_no_ii.text()).first():
                    session_chip_id.add(new_approval_no)
                    session_chip_id.commit()
                if not session_chip_id.query(ProductCategory).filter_by(
                        product_category=self.cb_prod_type_II.currentText()).first():
                    session_chip_id.add(new_product_category)
                    session_chip_id.commit()
                row = ChipId(chip_id=tup[0], asset_no=tup[1],
                             work_order_no_id=session_chip_id.query(WorkOrderNo).filter_by(
                                 work_order_no=self.value_order_II.text().strip().upper()).first().id,
                             approval_no_id=session_chip_id.query(ApprovalNo).filter_by(
                                 approval_no=self.le_approval_no_ii.text()).first().id,
                             product_category_id=session_chip_id.query(ProductCategory).filter_by(
                                 product_category=self.cb_prod_type_II.currentText()).first().id)
                session_chip_id.add(row)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = MyMainWindow()
    sys.exit(app.exec_())
